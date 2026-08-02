"""Excel input: the grid is given, so these tests are about what a cell *shows*.

The risk with a spreadsheet is not detection -- it is putting a number in the
output that appears nowhere on screen (0.291 instead of 29.1%), or cutting a
table in half at a row that only looks blank.
"""

import datetime
import zipfile

from docx import Document
from openpyxl import Workbook

from pdf_table_tool.pipeline import (
    SUPPORTED_SUFFIXES,
    PDFTableFlattenerPipeline,
    output_suffix_for,
)
from pdf_table_tool.xlsx_flattener import (
    XlsxTableFlattener,
    _cell_text,
    has_uncached_formulas,
    sheet_blocks,
)


def _book(tmp_path, build, name="book.xlsx"):
    workbook = Workbook()
    build(workbook)
    path = tmp_path / name
    workbook.save(path)
    return str(path)


def _paragraphs(path):
    return [p.text for p in Document(path).paragraphs if p.text.strip()]


# ------------------------------------------------------------------ cell text
def _formatted(value, number_format="General"):
    workbook = Workbook()
    cell = workbook.active.cell(row=1, column=1, value=value)
    cell.number_format = number_format
    return _cell_text(cell)


def test_a_percentage_is_read_as_the_sheet_shows_it():
    """Excel stores 29.1% as 0.291; the stored number is not what the reader sees."""
    assert _formatted(0.291, "0.0%") == "29.1%"
    assert _formatted(0.5, "0%") == "50%"


def test_a_grouped_number_keeps_its_thousands_separators():
    assert _formatted(1090350000, "#,##0") == "1,090,350,000"
    assert _formatted(1234.5, "#,##0.00") == "1,234.50"


def test_a_plain_number_does_not_gain_a_separator_it_never_had():
    assert _formatted(415) == "415"
    assert _formatted(3.5) == "3.5"


def test_dates_are_rendered_day_first():
    assert _formatted(datetime.datetime(2026, 6, 13)) == "13/06/2026"
    assert _formatted(datetime.datetime(2026, 6, 13, 8, 30)) == "13/06/2026 08:30"


def test_an_empty_cell_yields_no_text():
    assert _formatted(None) == ""


# --------------------------------------------------------------------- blocks
def test_a_blank_row_separates_two_tables():
    workbook = Workbook()
    sheet = workbook.active
    for row in (["A", "B"], ["1", "2"]):
        sheet.append(row)
    sheet.append([])
    for row in (["C", "D"], ["3", "4"]):
        sheet.append(row)
    assert sheet_blocks(sheet) == [(1, 2, 1, 2), (4, 5, 1, 2)]


def test_a_merged_cell_does_not_cut_its_own_table_in_half():
    """The rows under a merged label read as blank, but the table runs on."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Nhóm A"
    sheet.merge_cells("A1:A3")
    for row in range(1, 4):
        sheet.cell(row=row, column=2, value=f"giá trị {row}")
    assert sheet_blocks(sheet) == [(1, 3, 1, 2)]


# ------------------------------------------------------------------- end to end
def test_a_sheet_becomes_bullets_labelled_by_its_header_row(tmp_path):
    def build(workbook):
        sheet = workbook.active
        sheet.title = "Doanh Thu"
        sheet.append(["Mảng Dịch Vụ", "Số Lượng", "Doanh Thu", "Tỷ Lệ"])
        sheet.append(["Thay dầu động cơ", 415, 1090350000, 0.291])
        sheet.cell(row=2, column=3).number_format = "#,##0"
        sheet.cell(row=2, column=4).number_format = "0.0%"

    src = _book(tmp_path, build)
    out = str(tmp_path / "out.docx")
    XlsxTableFlattener(use_llm=False, verify_output=False).process(src, out)

    assert _paragraphs(out) == [
        "- Mảng Dịch Vụ: Thay dầu động cơ  |  Số Lượng: 415  |  "
        "Doanh Thu: 1,090,350,000  |  Tỷ Lệ: 29.1%"
    ]


def test_a_header_merged_over_sub_labels_names_both_of_its_columns(tmp_path):
    """Excel records the merge, so the two header rows read as one band.

    Without this the rows below carry no labels at all and the two header rows
    come out as bullets of their own.
    """
    def build(workbook):
        sheet = workbook.active
        sheet["A1"] = "Khách"
        sheet.merge_cells("A1:B1")
        sheet["C1"] = "Doanh thu"
        sheet.merge_cells("C1:D1")
        sheet["A2"], sheet["B2"] = "Họ tên", "Biển số"
        sheet["C2"], sheet["D2"] = "Trước thuế", "Tỷ lệ"
        sheet["A3"], sheet["B3"] = "Trần Văn Hải", "16F-242.23"
        sheet["C3"], sheet["D3"] = 1800000, 0.291
        sheet.cell(row=3, column=3).number_format = "#,##0"
        sheet.cell(row=3, column=4).number_format = "0.0%"

    out = str(tmp_path / "out.docx")
    XlsxTableFlattener(use_llm=False, verify_output=False).process(
        _book(tmp_path, build), out
    )
    assert _paragraphs(out) == [
        "- Khách Họ tên: Trần Văn Hải  |  Khách Biển số: 16F-242.23  |  "
        "Doanh thu Trước thuế: 1,800,000  |  Doanh thu Tỷ lệ: 29.1%"
    ]


def test_print_titles_declare_the_header_band(tmp_path):
    """"Rows to repeat at top" is Excel's "this row is the header"."""
    def build(workbook):
        sheet = workbook.active
        sheet.append(["Mã", "Ghi chú rất dài về dịch vụ đã thực hiện cho xe"])
        sheet.append(["X1", "Thay dầu"])
        sheet.print_title_rows = "$1:$1"

    out = str(tmp_path / "out.docx")
    XlsxTableFlattener(use_llm=False, verify_output=False).process(
        _book(tmp_path, build), out
    )
    assert _paragraphs(out) == [
        "- Mã: X1  |  Ghi chú rất dài về dịch vụ đã thực hiện cho xe: Thay dầu"
    ]


def test_every_sheet_is_named_when_a_workbook_has_more_than_one(tmp_path):
    def build(workbook):
        first = workbook.active
        first.title = "Một"
        first.append(["Tên", "Tuổi"])
        first.append(["Nam", 25])
        second = workbook.create_sheet("Hai")
        second.append(["Mã", "Giá"])
        second.append(["X1", 10])

    out = str(tmp_path / "out.docx")
    XlsxTableFlattener(use_llm=False, verify_output=False).process(
        _book(tmp_path, build), out
    )
    assert _paragraphs(out) == [
        "Một",
        "- Tên: Nam  |  Tuổi: 25",
        "Hai",
        "- Mã: X1  |  Giá: 10",
    ]


def test_a_single_sheet_workbook_is_not_captioned_with_sheet1(tmp_path):
    """A heading of "Sheet1" would add a word the workbook never contained."""
    def build(workbook):
        sheet = workbook.active
        sheet.append(["Tên", "Tuổi"])
        sheet.append(["Nam", 25])

    out = str(tmp_path / "out.docx")
    XlsxTableFlattener(use_llm=False, verify_output=False).process(
        _book(tmp_path, build), out
    )
    assert _paragraphs(out) == ["- Tên: Nam  |  Tuổi: 25"]


def test_nothing_a_cell_shows_is_lost(tmp_path):
    def build(workbook):
        sheet = workbook.active
        sheet["A1"] = "BÁO CÁO 2026"
        sheet.merge_cells("A1:C1")
        sheet.append([])
        sheet.append(["Mã HĐ", "Ngày Xuất", "Thành Tiền"])
        sheet.append(["HD-2026-0001", datetime.date(2026, 6, 13), 1800000])
        sheet.cell(row=4, column=3).number_format = "#,##0"

    src = _book(tmp_path, build)
    out = str(tmp_path / "out.docx")
    summary = XlsxTableFlattener(use_llm=False).process(src, out)
    assert summary["verification_passed"], summary["verification"].describe()
    assert summary["status"] == "success"
    # The merged title is a block of its own, not a table.
    assert summary["total_tables_flattened"] == 1


def test_the_pipeline_routes_excel_and_corrects_the_extension(tmp_path):
    def build(workbook):
        sheet = workbook.active
        sheet.append(["Tên", "Tuổi"])
        sheet.append(["Nam", 25])

    src = _book(tmp_path, build)
    assert output_suffix_for(src) == ".docx"
    assert ".xlsx" in SUPPORTED_SUFFIXES

    # A caller that asks for .xlsx must not get a Word document wearing it.
    summary = PDFTableFlattenerPipeline(use_llm=False).process(
        src, str(tmp_path / "out.xlsx")
    )
    assert summary["output_file"].endswith(".docx")
    assert summary["verification_passed"]


def test_a_formula_excel_never_calculated_is_reported(tmp_path):
    """The token check cannot see this: the cell is empty on both sides."""
    def build(workbook):
        sheet = workbook.active
        sheet.append(["Mục", "Giá", "Số lượng", "Thành tiền"])
        sheet.append(["A", 100, 3, "=B2*C2"])

    out = str(tmp_path / "out.docx")
    summary = XlsxTableFlattener(use_llm=False).process(_book(tmp_path, build), out)
    assert summary["uncached_formulas"] is True


def test_a_formula_excel_did_calculate_raises_no_warning(tmp_path):
    """Excel stores the result next to the formula; that file is fine as it is.

    openpyxl cannot write a cached value, so the sheet XML is rewritten the way
    Excel itself would have saved it.
    """
    def build(workbook):
        sheet = workbook.active
        sheet.append(["Mục", "Thành tiền"])
        sheet.append(["A", "=1*300"])

    src = _book(tmp_path, build)
    saved = str(tmp_path / "saved-by-excel.xlsx")
    with zipfile.ZipFile(src) as source, zipfile.ZipFile(saved, "w") as target:
        for item in source.namelist():
            blob = source.read(item)
            if item == "xl/worksheets/sheet1.xml":
                blob = blob.replace(b"<v></v>", b"<v>300</v>")
            target.writestr(item, blob)

    assert has_uncached_formulas(src) is True
    assert has_uncached_formulas(saved) is False


def test_a_workbook_without_formulas_raises_no_such_warning(tmp_path):
    def build(workbook):
        sheet = workbook.active
        sheet.append(["Mục", "Giá"])
        sheet.append(["A", 100])

    out = str(tmp_path / "out.docx")
    summary = XlsxTableFlattener(use_llm=False).process(_book(tmp_path, build), out)
    assert "uncached_formulas" not in summary


def test_a_pdf_still_keeps_its_own_extension():
    assert output_suffix_for("report.pdf") == ".pdf"
    assert output_suffix_for("report.docx") == ".docx"
